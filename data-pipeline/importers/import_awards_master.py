from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path

import mysql.connector
from openpyxl import load_workbook


DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "000000",
    "database": "teaching_awards",
    "charset": "utf8mb4",
    "autocommit": False,
}

MASTER_TABLE = "awards_master"

PROJECT_ROOT = Path(__file__).resolve().parents[2]
ZHEJIANG_CSV = PROJECT_ROOT / "data" / "processed" / "zhejiang" / "zhejiang_awards.csv"
PROVINCES_DIR = PROJECT_ROOT / "data" / "raw" / "provinces"
SHANGHAI_XLSX = PROVINCES_DIR / "2\u4e0a\u6d77\u5e02" / "\u4e0a\u6d77\u5e02\u6559\u5b66\u6210\u679c\u5956_527\u9879_\u603b\u8868.xlsx"
SHAANXI_XLSX = PROVINCES_DIR / "4\u9655\u897f\u7701" / "\u9655\u897f\u7701\u6559\u5b66\u6210\u679c\u5956_295\u9879_\u603b\u8868.xlsx"
HUBEI_XLSX = PROVINCES_DIR / "6\u6e56\u5317\u7701" / "\u6e56\u5317\u7701\u6559\u5b66\u6210\u679c\u5956_399\u9879_\u603b\u8868.xlsx"
HUNAN_XLSX = PROVINCES_DIR / "9\u6e56\u5357\u7701" / "\u6e56\u5357\u7701\u6559\u5b66\u6210\u679c\u5956_435\u9879_\u603b\u8868.xlsx"

COL_TITLE = "\u6210\u679c\u540d\u79f0"
COL_AUTHORS = "\u5b8c\u6210\u4eba"
COL_ORG = "\u5b8c\u6210\u5355\u4f4d"
COL_LEVEL = "\u83b7\u5956\u7b49\u7ea7"
COL_FIRST_ORG_LEVEL = "\u7b2c\u4e00\u5b8c\u6210\u5355\u4f4d\u5c42\u6b21"
COL_FIRST_ORG_TYPE = "\u7b2c\u4e00\u5b8c\u6210\u5355\u4f4d\u5b66\u79d1\u7c7b\u578b"
COL_FIRST_ORG_LOCATION = "\u7b2c\u4e00\u5b8c\u6210\u5355\u4f4d\u5730\u7406\u4f4d\u7f6e"
COL_AWARD_YEAR = "\u83b7\u5956\u5e74\u4efd"

PROVINCE_ZHEJIANG = "\u6d59\u6c5f\u7701"
PROVINCE_SHANGHAI = "\u4e0a\u6d77\u5e02"
PROVINCE_SHAANXI = "\u9655\u897f\u7701"
PROVINCE_HUBEI = "\u6e56\u5317\u7701"
PROVINCE_HUNAN = "\u6e56\u5357\u7701"


CREATE_TABLE_SQL = f"""
CREATE TABLE IF NOT EXISTS `{MASTER_TABLE}` (
    `id` BIGINT UNSIGNED NOT NULL AUTO_INCREMENT,
    `province` VARCHAR(32) NOT NULL,
    `award_year` INT NOT NULL,
    `award_level` VARCHAR(32) NULL,
    `award_title` VARCHAR(512) NOT NULL,
    `authors` TEXT NULL,
    `org_name` VARCHAR(255) NOT NULL,
    `first_org_level` VARCHAR(32) NULL,
    `first_org_type` VARCHAR(32) NULL,
    `first_org_location` VARCHAR(64) NULL,
    `source_file` VARCHAR(255) NULL,
    `source_sheet` VARCHAR(128) NULL,
    `source_row_no` INT NULL,
    `created_at` DATETIME NOT NULL,
    `updated_at` DATETIME NOT NULL,
    PRIMARY KEY (`id`),
    KEY `idx_awards_master_province_year` (`province`, `award_year`),
    KEY `idx_awards_master_province_year_level` (`province`, `award_year`, `award_level`),
    KEY `idx_awards_master_org_name` (`org_name`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
"""


INSERT_SQL = f"""
INSERT INTO `{MASTER_TABLE}` (
    province,
    award_year,
    award_level,
    award_title,
    authors,
    org_name,
    first_org_level,
    first_org_type,
    first_org_location,
    source_file,
    source_sheet,
    source_row_no,
    created_at,
    updated_at
) VALUES (
    %(province)s,
    %(award_year)s,
    %(award_level)s,
    %(award_title)s,
    %(authors)s,
    %(org_name)s,
    %(first_org_level)s,
    %(first_org_type)s,
    %(first_org_location)s,
    %(source_file)s,
    %(source_sheet)s,
    %(source_row_no)s,
    %(created_at)s,
    %(updated_at)s
)
"""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--create-only", action="store_true")
    parser.add_argument("--shanghai-year", type=int)
    parser.add_argument("--shaanxi-year", type=int)
    parser.add_argument("--hubei-year", type=int)
    parser.add_argument("--hunan-year", type=int)
    parser.add_argument("--include-zhejiang", dest="include_zhejiang", action="store_true")
    parser.add_argument("--skip-zhejiang", dest="include_zhejiang", action="store_false")
    parser.add_argument("--skip-shanghai", action="store_true")
    parser.add_argument("--skip-shaanxi", action="store_true")
    parser.add_argument("--skip-hubei", action="store_true")
    parser.add_argument("--skip-hunan", action="store_true")
    parser.set_defaults(include_zhejiang=False)
    return parser.parse_args()


def clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def clean_year(value) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return int(float(text))


def build_row(
    *,
    province: str,
    award_year: int,
    award_level: str | None,
    award_title: str,
    authors: str | None,
    org_name: str,
    first_org_level: str | None,
    first_org_type: str | None,
    first_org_location: str | None,
    source_file: str,
    source_sheet: str | None,
    source_row_no: int,
    now: datetime,
) -> dict:
    return {
        "province": province,
        "award_year": award_year,
        "award_level": award_level,
        "award_title": award_title,
        "authors": authors,
        "org_name": org_name,
        "first_org_level": first_org_level,
        "first_org_type": first_org_type,
        "first_org_location": first_org_location,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_row_no": source_row_no,
        "created_at": now,
        "updated_at": now,
    }


def load_zhejiang_rows(now: datetime) -> list[dict]:
    rows: list[dict] = []
    last_error = None
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            with ZHEJIANG_CSV.open("r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                for idx, row in enumerate(reader, start=2):
                    award_title = clean_text(row.get(COL_TITLE))
                    org_name = clean_text(row.get(COL_ORG))
                    award_year = clean_year(row.get(COL_AWARD_YEAR))
                    if not award_title or not org_name or award_year is None:
                        continue
                    rows.append(
                        build_row(
                            province=PROVINCE_ZHEJIANG,
                            award_year=award_year,
                            award_level=None,
                            award_title=award_title,
                            authors=clean_text(row.get(COL_AUTHORS)),
                            org_name=org_name,
                            first_org_level=clean_text(row.get(COL_FIRST_ORG_LEVEL)),
                            first_org_type=clean_text(row.get(COL_FIRST_ORG_TYPE)),
                            first_org_location=clean_text(row.get(COL_FIRST_ORG_LOCATION)),
                            source_file=ZHEJIANG_CSV.name,
                            source_sheet=None,
                            source_row_no=idx,
                            now=now,
                        )
                    )
            return rows
        except UnicodeDecodeError as exc:
            rows.clear()
            last_error = exc
            continue
    if last_error:
        raise last_error
    return rows


def load_province_excel_rows(now: datetime, xlsx_path: Path, province: str, year: int) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = [clean_text(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    pos = {name: idx for idx, name in enumerate(header)}

    required = [COL_TITLE, COL_AUTHORS, COL_ORG, COL_LEVEL]
    missing = [name for name in required if name not in pos]
    if missing:
        raise ValueError(f"{province} sheet is missing columns: {missing}")

    rows: list[dict] = []
    for row_no, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        award_title = clean_text(values[pos[COL_TITLE]])
        org_name = clean_text(values[pos[COL_ORG]])
        if not award_title or not org_name:
            continue
        rows.append(
            build_row(
                province=province,
                award_year=year,
                award_level=clean_text(values[pos[COL_LEVEL]]),
                award_title=award_title,
                authors=clean_text(values[pos[COL_AUTHORS]]),
                org_name=org_name,
                first_org_level=clean_text(values[pos.get(COL_FIRST_ORG_LEVEL)]) if COL_FIRST_ORG_LEVEL in pos else None,
                first_org_type=clean_text(values[pos.get(COL_FIRST_ORG_TYPE)]) if COL_FIRST_ORG_TYPE in pos else None,
                first_org_location=clean_text(values[pos.get(COL_FIRST_ORG_LOCATION)]) if COL_FIRST_ORG_LOCATION in pos else None,
                source_file=xlsx_path.name,
                source_sheet=ws.title,
                source_row_no=row_no,
                now=now,
            )
        )
    return rows


def main() -> None:
    args = parse_args()
    now = datetime.now()

    db = mysql.connector.connect(**DB_CONFIG)
    cur = db.cursor()

    try:
        cur.execute(CREATE_TABLE_SQL)
        db.commit()
        print(f"CREATED_TABLE: {MASTER_TABLE}")

        if args.create_only:
            return

        batch_rows: list[dict] = []
        imported_provinces: list[str] = []

        if args.include_zhejiang:
            zhejiang_rows = load_zhejiang_rows(now)
            batch_rows.extend(zhejiang_rows)
            imported_provinces.append(PROVINCE_ZHEJIANG)
            print(f"ZHEJIANG_ROWS: {len(zhejiang_rows)}")

        if not args.skip_shanghai:
            if not args.shanghai_year:
                raise ValueError("--shanghai-year is required unless --skip-shanghai is set")
            shanghai_rows = load_province_excel_rows(now, SHANGHAI_XLSX, PROVINCE_SHANGHAI, args.shanghai_year)
            batch_rows.extend(shanghai_rows)
            imported_provinces.append(PROVINCE_SHANGHAI)
            print(f"SHANGHAI_ROWS: {len(shanghai_rows)}")

        if not args.skip_shaanxi:
            if not args.shaanxi_year:
                raise ValueError("--shaanxi-year is required unless --skip-shaanxi is set")
            shaanxi_rows = load_province_excel_rows(now, SHAANXI_XLSX, PROVINCE_SHAANXI, args.shaanxi_year)
            batch_rows.extend(shaanxi_rows)
            imported_provinces.append(PROVINCE_SHAANXI)
            print(f"SHAANXI_ROWS: {len(shaanxi_rows)}")

        if not args.skip_hubei:
            if not args.hubei_year:
                raise ValueError("--hubei-year is required unless --skip-hubei is set")
            hubei_rows = load_province_excel_rows(now, HUBEI_XLSX, PROVINCE_HUBEI, args.hubei_year)
            batch_rows.extend(hubei_rows)
            imported_provinces.append(PROVINCE_HUBEI)
            print(f"HUBEI_ROWS: {len(hubei_rows)}")

        if not args.skip_hunan:
            if not args.hunan_year:
                raise ValueError("--hunan-year is required unless --skip-hunan is set")
            hunan_rows = load_province_excel_rows(now, HUNAN_XLSX, PROVINCE_HUNAN, args.hunan_year)
            batch_rows.extend(hunan_rows)
            imported_provinces.append(PROVINCE_HUNAN)
            print(f"HUNAN_ROWS: {len(hunan_rows)}")

        for province in imported_provinces:
            cur.execute(f"DELETE FROM `{MASTER_TABLE}` WHERE province = %s", (province,))

        if batch_rows:
            cur.executemany(INSERT_SQL, batch_rows)

        db.commit()
        print(f"IMPORTED_TOTAL: {len(batch_rows)}")
        print(f"IMPORTED_PROVINCES: {imported_provinces}")
    except Exception:
        db.rollback()
        raise
    finally:
        cur.close()
        db.close()


if __name__ == "__main__":
    main()
