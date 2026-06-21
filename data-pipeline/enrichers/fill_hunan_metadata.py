from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook


LEVEL_985 = {
    "中南大学",
    "湖南大学",
    "国防科技大学",
}

LEVEL_211 = {
    "湖南师范大学",
}

TYPE_MAP = {
    "中南大学": "综合类",
    "中南林业科技大学": "农林类",
    "吉首大学": "综合类",
    "国防科技大学": "理工类",
    "张家界学院": "综合类",
    "怀化学院": "综合类",
    "湖南中医药大学": "医药类",
    "湖南交通工程学院": "理工类",
    "湖南人文科技学院": "综合类",
    "湖南信息学院": "理工类",
    "湖南农业大学": "农林类",
    "湖南医药学院": "医药类",
    "湖南城市学院": "综合类",
    "湖南大学": "综合类",
    "湖南工业大学": "理工类",
    "湖南工商大学": "财经类",
    "湖南工学院": "理工类",
    "湖南工程学院": "理工类",
    "湖南师范大学": "师范类",
    "湖南文理学院": "综合类",
    "湖南文理学院芙蓉学院": "综合类",
    "湖南涉外经济学院": "财经类",
    "湖南理工学院": "理工类",
    "湖南省教科院": "其他类",
    "湖南第一师范学院": "师范类",
    "湖南警察学院": "政法类",
    "湖南财政经济学院": "财经类",
    "湘南学院": "综合类",
    "湘潭大学": "综合类",
    "湘潭理工学院": "理工类",
    "衡阳师范学院": "师范类",
    "邵阳学院": "综合类",
    "长沙医学院": "医药类",
    "长沙学院": "综合类",
    "长沙师范学院": "师范类",
    "长沙理工大学": "理工类",
}

CITY_MAP = {
    "中南大学": "长沙市",
    "中南林业科技大学": "长沙市",
    "南华大学": "衡阳市",
    "吉首大学": "吉首市",
    "国防科技大学": "长沙市",
    "张家界学院": "张家界市",
    "怀化学院": "怀化市",
    "湖南中医药大学": "长沙市",
    "湖南交通工程学院": "衡阳市",
    "湖南人文科技学院": "娄底市",
    "湖南信息学院": "长沙市",
    "湖南农业大学": "长沙市",
    "湖南医药学院": "怀化市",
    "湖南城市学院": "益阳市",
    "湖南大学": "长沙市",
    "湖南女子学院": "长沙市",
    "湖南工业大学": "株洲市",
    "湖南工商大学": "长沙市",
    "湖南工学院": "衡阳市",
    "湖南工程学院": "湘潭市",
    "湖南师范大学": "长沙市",
    "湖南应用技术学院": "常德市",
    "湖南文理学院": "常德市",
    "湖南文理学院芙蓉学院": "常德市",
    "湖南涉外经济学院": "长沙市",
    "湖南理工学院": "岳阳市",
    "湖南省教科院": "长沙市",
    "湖南科技大学": "湘潭市",
    "湖南科技学院": "永州市",
    "湖南第一师范学院": "长沙市",
    "湖南警察学院": "长沙市",
    "湖南财政经济学院": "长沙市",
    "湘南学院": "郴州市",
    "湘潭大学": "湘潭市",
    "湘潭理工学院": "湘潭市",
    "衡阳师范学院": "衡阳市",
    "邵阳学院": "邵阳市",
    "长沙医学院": "长沙市",
    "长沙学院": "长沙市",
    "长沙师范学院": "长沙市",
    "长沙理工大学": "长沙市",
}

FIRST_UNIT_PREFIXES = sorted({*LEVEL_985, *LEVEL_211, *TYPE_MAP.keys(), *CITY_MAP.keys()}, key=len, reverse=True)

COL_COMPLETION_UNIT = "完成单位"
COL_LEVEL = "第一完成单位层次"
COL_TYPE = "第一完成单位学科类型"
COL_CITY = "第一完成单位地理位置"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--in-place", action="store_true")
    return parser.parse_args()


def ensure_columns(ws) -> dict[str, int]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    existing = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for name in [COL_LEVEL, COL_TYPE, COL_CITY]:
        if name not in existing:
            ws.insert_cols(ws.max_column)
            ws.cell(1, ws.max_column).value = name
            existing = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    return existing


def clean_first_unit(raw: str) -> str:
    text = str(raw or "").strip().replace(" ", "")
    if not text:
        return ""
    for prefix in FIRST_UNIT_PREFIXES:
        if text.startswith(prefix):
            return prefix
    text = re.split(r"[、，,；;]+", text)[0].strip()
    for prefix in FIRST_UNIT_PREFIXES:
        if text.startswith(prefix):
            return prefix
    return text


def level_of(unit: str) -> str:
    if unit in LEVEL_985:
        return "985"
    if unit in LEVEL_211:
        return "211"
    if unit:
        return "其他"
    return ""


def main() -> None:
    args = parse_args()
    src = args.source
    if not src.exists():
        raise FileNotFoundError(src)

    out = src if args.in_place else src.with_name(f"{src.stem}_补充字段{src.suffix}")

    wb = load_workbook(src)
    ws = wb[wb.sheetnames[0]]
    header_pos = ensure_columns(ws)

    unresolved_type: set[str] = set()
    unresolved_city: set[str] = set()

    for row in range(2, ws.max_row + 1):
        raw_unit = str(ws.cell(row, header_pos[COL_COMPLETION_UNIT]).value or "").strip()
        if not raw_unit:
            continue

        first_unit = clean_first_unit(raw_unit)
        ws.cell(row, header_pos[COL_LEVEL]).value = level_of(first_unit)

        school_type = TYPE_MAP.get(first_unit, "")
        city = CITY_MAP.get(first_unit, "")

        ws.cell(row, header_pos[COL_TYPE]).value = school_type
        ws.cell(row, header_pos[COL_CITY]).value = city

        if first_unit and not school_type:
            unresolved_type.add(first_unit)
        if first_unit and not city:
            unresolved_city.add(first_unit)

    wb.save(out)
    print(f"OUTPUT: {out}")
    print(f"UNRESOLVED_TYPE: {sorted(unresolved_type)}")
    print(f"UNRESOLVED_CITY: {sorted(unresolved_city)}")


if __name__ == "__main__":
    main()
