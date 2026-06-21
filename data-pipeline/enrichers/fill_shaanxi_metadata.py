from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


LEVEL_985 = {
    "西安交通大学",
    "西北工业大学",
    "西北农林科技大学",
    "国防科技大学",
}

LEVEL_211 = {
    "西安电子科技大学",
    "陕西师范大学",
    "长安大学",
    "西北大学",
    "空军军医大学",
}

TYPE_MAP = {
    "西安交通大学": "综合类",
    "西北工业大学": "理工类",
    "西北农林科技大学": "农林类",
    "西安电子科技大学": "理工类",
    "陕西师范大学": "师范类",
    "长安大学": "理工类",
    "西北大学": "综合类",
    "西安理工大学": "理工类",
    "西安建筑科技大学": "理工类",
    "陕西科技大学": "理工类",
    "西安科技大学": "理工类",
    "西安石油大学": "理工类",
    "延安大学": "综合类",
    "西安工业大学": "理工类",
    "西安工程大学": "理工类",
    "西安外国语大学": "语言类",
    "西北政法大学": "政法类",
    "西安邮电大学": "理工类",
    "西安美术学院": "艺术类",
    "西安医学院": "医药类",
    "榆林学院": "其他类",
    "西安航空学院": "理工类",
    "西京学院": "理工类",
    "咸阳师范学院": "师范类",
    "西安培华学院": "综合类",
    "西安外事学院": "综合类",
    "西安思源学院": "综合类",
    "西安文理学院": "综合类",
    "西安欧亚学院": "其他类",
    "西安翻译学院": "语言类",
    "西藏民族大学": "综合类",
    "陕西国际商贸学院": "财经类",
    "陕西学前师范学院": "师范类",
    "陕西开放大学": "其他类",
    "陕西服装工程学院": "其他类",
    "陕西理工大学": "理工类",
    "空军军医大学": "医药类",
    "火箭军工程大学": "理工类",
    "国防科技大学": "理工类",
    "商洛学院": "综合类",
    "安康学院": "综合类",
    "宝鸡文理学院": "综合类",
    "武警工程大学": "理工类",
    "渭南师范学院": "师范类",
    "空军工程大学": "理工类",
    "中国人民解放军空军工程大学": "理工类",
    "西安财经大学": "财经类",
    "西安音乐学院": "艺术类",
    "西安体育学院": "体育类",
    "陕西中医药大学": "医药类",
}

CITY_MAP = {
    "西安交通大学": "西安市",
    "西北工业大学": "西安市",
    "西北农林科技大学": "咸阳市",
    "西安电子科技大学": "西安市",
    "陕西师范大学": "西安市",
    "长安大学": "西安市",
    "西北大学": "西安市",
    "西安理工大学": "西安市",
    "西安建筑科技大学": "西安市",
    "陕西科技大学": "西安市",
    "西安科技大学": "西安市",
    "西安石油大学": "西安市",
    "延安大学": "延安市",
    "西安工业大学": "西安市",
    "西安工程大学": "西安市",
    "西安外国语大学": "西安市",
    "西北政法大学": "西安市",
    "西安邮电大学": "西安市",
    "西安美术学院": "西安市",
    "西安医学院": "西安市",
    "榆林学院": "榆林市",
    "西安航空学院": "西安市",
    "西京学院": "西安市",
    "咸阳师范学院": "咸阳市",
    "西安培华学院": "西安市",
    "西安外事学院": "西安市",
    "西安思源学院": "西安市",
    "西安文理学院": "西安市",
    "西安欧亚学院": "西安市",
    "西安翻译学院": "西安市",
    "西藏民族大学": "咸阳市",
    "陕西国际商贸学院": "咸阳市",
    "陕西学前师范学院": "西安市",
    "陕西开放大学": "西安市",
    "陕西服装工程学院": "咸阳市",
    "陕西理工大学": "汉中市",
    "空军军医大学": "西安市",
    "火箭军工程大学": "西安市",
    "国防科技大学": "长沙市",
    "商洛学院": "商洛市",
    "安康学院": "安康市",
    "宝鸡文理学院": "宝鸡市",
    "武警工程大学": "西安市",
    "渭南师范学院": "渭南市",
    "空军工程大学": "西安市",
    "中国人民解放军空军工程大学": "西安市",
    "西安财经大学": "西安市",
    "西安音乐学院": "西安市",
    "西安体育学院": "西安市",
    "陕西中医药大学": "咸阳市",
}

FIRST_UNIT_PREFIXES = sorted(
    {
        *LEVEL_985,
        *LEVEL_211,
        *TYPE_MAP.keys(),
        *CITY_MAP.keys(),
    },
    key=len,
    reverse=True,
)

COL_COMPLETION_UNIT = "完成单位"
LEGACY_COL_FIRST_AUTHOR = "\u7b2c\u4e00\u5b8c\u6210\u4eba\u8eab\u4efd"
COL_LEVEL = "第一完成单位层次"
COL_TYPE = "第一完成单位学科类型"
COL_CITY = "第一完成单位地理位置"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--in-place", action="store_true")
    return parser.parse_args()


def clean_first_unit(raw: str) -> str:
    text = str(raw or "").strip().replace(" ", "")
    if not text:
        return ""
    for prefix in FIRST_UNIT_PREFIXES:
        if text.startswith(prefix):
            return prefix
    for sep in ["、", "，", ",", "；", ";"]:
        if sep in text:
            text = text.split(sep)[0].strip()
            break
    for prefix in FIRST_UNIT_PREFIXES:
        if text.startswith(prefix):
            return prefix
    return text


def level_of(unit: str) -> str:
    if unit in LEVEL_985:
        return "985"
    if unit in LEVEL_211:
        return "211"
    return "其他"


def ensure_output_columns(ws) -> dict[str, int]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if LEGACY_COL_FIRST_AUTHOR in headers:
        ws.delete_cols(headers.index(LEGACY_COL_FIRST_AUTHOR) + 1)

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    unit_idx = headers.index(COL_COMPLETION_UNIT) + 1
    existing = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    wanted = [COL_LEVEL, COL_TYPE, COL_CITY]
    for name in wanted:
        if name not in existing:
            insert_at = unit_idx + 1
            ws.insert_cols(insert_at)
            ws.cell(1, insert_at).value = name
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            existing = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            unit_idx = headers.index(COL_COMPLETION_UNIT) + 1

    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def main() -> None:
    args = parse_args()
    src = args.source
    if not src.exists():
        raise FileNotFoundError(src)

    out = src if args.in_place else src.with_name(f"{src.stem}_补充字段{src.suffix}")

    wb = load_workbook(src)
    ws = wb[wb.sheetnames[0]]
    header_pos = ensure_output_columns(ws)

    unresolved_city: set[str] = set()
    unresolved_unit: set[str] = set()

    for row in range(2, ws.max_row + 1):
        raw_unit = str(ws.cell(row, header_pos[COL_COMPLETION_UNIT]).value or "").strip()
        if not raw_unit:
            continue

        first_unit = clean_first_unit(raw_unit)
        if not first_unit:
            unresolved_unit.add(raw_unit)
            continue

        level = level_of(first_unit)
        school_type = TYPE_MAP.get(first_unit, "其他类")
        city = CITY_MAP.get(first_unit, "")

        ws.cell(row, header_pos[COL_LEVEL]).value = level
        ws.cell(row, header_pos[COL_TYPE]).value = school_type
        ws.cell(row, header_pos[COL_CITY]).value = city

        if not city:
            unresolved_city.add(first_unit)

    wb.save(out)

    print(f"OUTPUT: {out}")
    print(f"UNRESOLVED_UNIT: {sorted(unresolved_unit)}")
    print(f"UNRESOLVED_CITY: {sorted(unresolved_city)}")


if __name__ == "__main__":
    main()
