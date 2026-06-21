from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


OUTPUT_NAME = "上海市教学成果奖_527项_总表_补充字段.xlsx"

LEVEL_MAP = {
    "复旦大学": "985",
    "同济大学": "985",
    "上海交通大学": "985",
    "华东师范大学": "985",
    "华东理工大学": "211",
    "东华大学": "211",
    "上海外国语大学": "211",
    "上海财经大学": "211",
    "上海大学": "211",
    "海军军医大学": "211",
    "上海中医药大学": "其他",
    "上海体育学院": "其他",
    "上海健康医学院": "其他",
    "上海公安学院": "其他",
    "上海商学院": "其他",
    "上海国家会计学院": "其他",
    "上海外国语大学贤达经济人文学院": "其他",
    "上海科技大学": "其他",
    "上海纽约大学": "其他",
    "上海对外经贸大学": "其他",
    "上海立信会计金融学院": "其他",
    "华东政法大学": "其他",
    "上海政法学院": "其他",
    "上海师范大学": "其他",
    "上海师范大学天华学院": "其他",
    "上海市教育科学研究院": "其他",
    "上海工程技术大学": "其他",
    "上海应用技术大学": "其他",
    "上海电力大学": "其他",
    "上海海事大学": "其他",
    "上海理工大学": "其他",
    "上海海洋大学": "其他",
    "上海电机学院": "其他",
    "上海第二工业大学": "其他",
    "上海建桥学院": "其他",
    "上海杉达学院": "其他",
    "上海戏剧学院": "其他",
    "上海视觉艺术学院": "其他",
    "上海音乐学院": "其他",
    "上海海关学院": "其他",
    "上海社会科学院": "其他",
    "上海市教育考试院": "其他",
    "上海市教育评估院": "其他",
    "中共上海市委党校": "其他",
}

TYPE_MAP = {
    "复旦大学": "综合类",
    "同济大学": "综合类",
    "上海交通大学": "综合类",
    "华东师范大学": "师范类",
    "华东理工大学": "理工类",
    "东华大学": "理工类",
    "上海大学": "综合类",
    "上海财经大学": "财经类",
    "上海中医药大学": "医药类",
    "上海健康医学院": "医药类",
    "海军军医大学": "医药类",
    "上海海洋大学": "农林类",
    "上海对外经贸大学": "财经类",
    "上海立信会计金融学院": "财经类",
    "上海商学院": "财经类",
    "上海国家会计学院": "财经类",
    "上海海关学院": "财经类",
    "上海外国语大学贤达经济人文学院": "财经类",
    "上海工程技术大学": "理工类",
    "上海应用技术大学": "理工类",
    "上海理工大学": "理工类",
    "上海电力大学": "理工类",
    "上海电机学院": "理工类",
    "上海第二工业大学": "理工类",
    "上海海事大学": "理工类",
    "上海科技大学": "理工类",
    "上海师范大学": "师范类",
    "上海师范大学天华学院": "师范类",
    "上海纽约大学": "综合类",
    "上海建桥学院": "综合类",
    "上海杉达学院": "综合类",
    "上海外国语大学": "其他类",
    "华东政法大学": "其他类",
    "上海政法学院": "其他类",
    "上海公安学院": "其他类",
    "上海戏剧学院": "其他类",
    "上海视觉艺术学院": "其他类",
    "上海音乐学院": "其他类",
    "上海体育学院": "其他类",
    "上海市教育科学研究院": "其他类",
    "上海市教育考试院": "其他类",
    "上海市教育评估院": "其他类",
    "上海社会科学院": "其他类",
    "中共上海市委党校": "其他类",
}

NEW_COLUMNS = [
    "第一完成单位层次",
    "第一完成单位学科类型",
    "第一完成单位地理位置",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--in-place", action="store_true")
    return parser.parse_args()


def ensure_columns(ws) -> dict[str, int]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    unit_idx = headers.index("完成单位") + 1

    existing = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    if all(name in existing for name in NEW_COLUMNS):
        return existing

    insert_at = unit_idx + 1
    for offset, name in enumerate(NEW_COLUMNS):
        ws.insert_cols(insert_at + offset)
        ws.cell(1, insert_at + offset).value = name

    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def main() -> None:
    args = parse_args()
    src = args.source
    if not src.exists():
        raise FileNotFoundError(src)

    out = src if args.in_place else src.with_name(OUTPUT_NAME)

    wb = load_workbook(src)
    ws = wb[wb.sheetnames[0]]
    header_pos = ensure_columns(ws)

    unresolved_level: set[str] = set()
    unresolved_type: set[str] = set()

    for row in range(2, ws.max_row + 1):
        unit = str(ws.cell(row, header_pos["完成单位"]).value or "").strip()
        if not unit:
            continue

        level = LEVEL_MAP.get(unit, "")
        school_type = TYPE_MAP.get(unit, "")

        ws.cell(row, header_pos["第一完成单位层次"]).value = level
        ws.cell(row, header_pos["第一完成单位学科类型"]).value = school_type
        ws.cell(row, header_pos["第一完成单位地理位置"]).value = ""

        if not level:
            unresolved_level.add(unit)
        if not school_type:
            unresolved_type.add(unit)

    wb.save(out)

    print(f"OUTPUT: {out}")
    print(f"UNRESOLVED_LEVEL: {sorted(unresolved_level)}")
    print(f"UNRESOLVED_TYPE: {sorted(unresolved_type)}")


if __name__ == "__main__":
    main()
