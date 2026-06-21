from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


TYPE_MAP = {
    "华中科技大学": "理工类",
    "武汉大学": "综合类",
    "中国地质大学（武汉）": "理工类",
    "华中农业大学": "农林类",
    "华中师范大学": "师范类",
    "武汉理工大学": "理工类",
    "武汉科技大学": "理工类",
    "湖北大学": "综合类",
    "中南财经政法大学": "财经类",
    "中南民族大学": "综合类",
    "武汉工程大学": "理工类",
    "武汉纺织大学": "理工类",
    "武汉轻工大学": "理工类",
    "武汉体育学院": "体育类",
    "武汉音乐学院": "艺术类",
    "湖北工业大学": "理工类",
    "湖北经济学院": "财经类",
    "湖北经济学院法商学院": "财经类",
    "湖北中医药大学": "医药类",
    "湖北医药学院": "医药类",
    "湖北医药学院药护学院": "医药类",
    "湖北工程学院": "综合类",
    "湖北师范大学": "师范类",
    "湖北第二师范学院": "师范类",
    "湖北理工学院": "理工类",
    "湖北科技学院": "理工类",
    "湖北美术学院": "艺术类",
    "湖北民族大学": "综合类",
    "湖北文理学院": "综合类",
    "湖北汽车工业学院": "理工类",
    "湖北汽车工业学院科技学院": "理工类",
    "湖北警官学院": "政法类",
    "湖北开放大学": "其他类",
    "湖北教育数字化研究院": "其他类",
    "长江大学": "综合类",
    "长江大学文理学院": "综合类",
    "三峡大学": "综合类",
    "江汉大学": "综合类",
    "海军工程大学": "理工类",
    "空军预警学院": "理工类",
    "信息支援部队工程大学": "理工类",
    "黄冈师范学院": "师范类",
    "汉江师范学院": "师范类",
    "荆楚理工学院": "理工类",
    "荆州学院": "综合类",
    "文华学院": "综合类",
    "武昌首义学院": "综合类",
    "武昌工学院": "理工类",
    "武昌理工学院": "理工类",
    "武汉东湖学院": "综合类",
    "武汉传媒学院": "艺术类",
    "武汉华夏理工学院": "理工类",
    "武汉学院": "综合类",
    "武汉工商学院": "财经类",
    "武汉工程科技学院": "理工类",
    "武汉文理学院": "综合类",
    "武汉晴川学院": "综合类",
    "武汉城市学院": "综合类",
    "武汉设计工程学院": "艺术类",
    "武汉商学院": "财经类",
    "武汉生物工程学院": "理工类",
    "湖北恩施学院": "综合类",
    "湖北商贸学院": "财经类",
    "汉口学院": "综合类",
}

CITY_MAP = {
    "华中科技大学": "武汉市",
    "武汉大学": "武汉市",
    "中国地质大学（武汉）": "武汉市",
    "华中农业大学": "武汉市",
    "华中师范大学": "武汉市",
    "武汉理工大学": "武汉市",
    "武汉科技大学": "武汉市",
    "湖北大学": "武汉市",
    "中南财经政法大学": "武汉市",
    "中南民族大学": "武汉市",
    "武汉工程大学": "武汉市",
    "武汉纺织大学": "武汉市",
    "武汉轻工大学": "武汉市",
    "武汉体育学院": "武汉市",
    "武汉音乐学院": "武汉市",
    "湖北工业大学": "武汉市",
    "湖北经济学院": "武汉市",
    "湖北经济学院法商学院": "武汉市",
    "湖北中医药大学": "武汉市",
    "湖北医药学院": "十堰市",
    "湖北医药学院药护学院": "十堰市",
    "湖北工程学院": "孝感市",
    "湖北师范大学": "黄石市",
    "湖北第二师范学院": "武汉市",
    "湖北理工学院": "黄石市",
    "湖北科技学院": "咸宁市",
    "湖北美术学院": "武汉市",
    "湖北民族大学": "恩施市",
    "湖北文理学院": "襄阳市",
    "湖北汽车工业学院": "十堰市",
    "湖北汽车工业学院科技学院": "十堰市",
    "湖北警官学院": "武汉市",
    "湖北开放大学": "武汉市",
    "湖北教育数字化研究院": "武汉市",
    "长江大学": "荆州市",
    "长江大学文理学院": "荆州市",
    "三峡大学": "宜昌市",
    "江汉大学": "武汉市",
    "海军工程大学": "武汉市",
    "空军预警学院": "武汉市",
    "信息支援部队工程大学": "武汉市",
    "黄冈师范学院": "黄冈市",
    "汉江师范学院": "十堰市",
    "荆楚理工学院": "荆门市",
    "荆州学院": "荆州市",
    "文华学院": "武汉市",
    "武昌首义学院": "武汉市",
    "武昌工学院": "武汉市",
    "武昌理工学院": "武汉市",
    "武汉东湖学院": "武汉市",
    "武汉传媒学院": "武汉市",
    "武汉华夏理工学院": "武汉市",
    "武汉学院": "武汉市",
    "武汉工商学院": "武汉市",
    "武汉工程科技学院": "武汉市",
    "武汉文理学院": "武汉市",
    "武汉晴川学院": "武汉市",
    "武汉城市学院": "武汉市",
    "武汉设计工程学院": "武汉市",
    "武汉商学院": "武汉市",
    "武汉生物工程学院": "武汉市",
    "湖北恩施学院": "恩施市",
    "湖北商贸学院": "武汉市",
    "汉口学院": "武汉市",
}

KNOWN_UNITS = sorted({*TYPE_MAP.keys(), *CITY_MAP.keys()}, key=len, reverse=True)

ROW_FIXES = {
    218: {
        "authors_append": "李晓艳、赵玲、RU GAO",
        "org": "华中科技大学",
    },
    253: {
        "authors_append": "袁龙、汪汝武、林昕、玛依努尔·买地亚尔",
        "org": "武汉科技大学",
    },
    316: {
        "authors_replace": "郭顺峰、姚炎昕、夏小林、李光、田萍、邓易",
        "org": "汉江师范学院",
    },
    386: {
        "authors_append": "林辉、卢刚、Robert M.Larkin",
        "org": "华中农业大学",
    },
    389: {
        "authors_append": "吴茜",
        "org": "湖北工业大学",
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--in-place", action="store_true")
    return parser.parse_args()


def ensure_columns(ws) -> dict[str, int]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    existing = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    if "第一完成单位学科类型" not in existing:
        ws.insert_cols(6)
        ws.cell(1, 6).value = "第一完成单位学科类型"
    if "第一完成单位地理位置" not in existing:
        ws.insert_cols(7)
        ws.cell(1, 7).value = "第一完成单位地理位置"

    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def clean_first_unit(raw: str) -> str:
    text = str(raw or "").strip().replace(" ", "")
    if not text:
        return ""
    for unit in KNOWN_UNITS:
        if text.startswith(unit):
            return unit
    for unit in KNOWN_UNITS:
        idx = text.find(unit)
        if idx != -1:
            return unit
    for sep in ["、", "，", ",", ";", "；"]:
        if sep in text:
            return text.split(sep)[0].strip()
    return text


def append_authors(existing: str, extra: str) -> str:
    base = (existing or "").strip()
    if base.endswith("、"):
        base = base[:-1]
    if not base:
        return extra
    return f"{base}、{extra}"


def apply_row_fixes(ws) -> None:
    for row, payload in ROW_FIXES.items():
        if "authors_replace" in payload:
            ws.cell(row, 3).value = payload["authors_replace"]
        else:
            current = str(ws.cell(row, 3).value or "")
            ws.cell(row, 3).value = append_authors(current, payload["authors_append"])
        ws.cell(row, 4).value = payload["org"]


def main() -> None:
    args = parse_args()
    src = args.source
    if not src.exists():
        raise FileNotFoundError(src)

    out = src if args.in_place else src.with_name(f"{src.stem}_补充字段{src.suffix}")
    wb = load_workbook(src)
    ws = wb[wb.sheetnames[0]]
    header_pos = ensure_columns(ws)
    apply_row_fixes(ws)

    unresolved_type: set[str] = set()
    unresolved_city: set[str] = set()

    for row in range(2, ws.max_row + 1):
        raw_unit = str(ws.cell(row, header_pos["完成单位"]).value or "").strip()
        if not raw_unit:
            continue
        first_unit = clean_first_unit(raw_unit)
        school_type = TYPE_MAP.get(first_unit, "其他类")
        city = CITY_MAP.get(first_unit, "")

        ws.cell(row, header_pos["第一完成单位学科类型"]).value = school_type
        ws.cell(row, header_pos["第一完成单位地理位置"]).value = city

        if first_unit not in TYPE_MAP:
            unresolved_type.add(first_unit)
        if not city:
            unresolved_city.add(first_unit)

    wb.save(out)
    print(f"OUTPUT: {out}")
    print(f"UNRESOLVED_TYPE: {sorted(unresolved_type)}")
    print(f"UNRESOLVED_CITY: {sorted(unresolved_city)}")


if __name__ == "__main__":
    main()
