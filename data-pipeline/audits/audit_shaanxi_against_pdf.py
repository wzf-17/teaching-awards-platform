from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook


SERIAL_RE = re.compile(r"^(\d+)\b(?:\s+(.*))?$")
TITLE_KEYS = (
    "培养",
    "实践",
    "体系",
    "模式",
    "构建",
    "创新",
    "导向",
    "协同",
    "赋能",
    "育人",
    "探索",
    "改革",
)
ORG_HINTS = (
    "大学",
    "学院",
    "研究院",
    "研究所",
    "中心",
    "基地",
    "学校",
    "公司",
    "医院",
)
SKIP_PREFIXES = ("附件", "序号 学校名称 申报成果名称 成果主要完成人姓名")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--undergrad-txt", type=Path, required=True)
    parser.add_argument("--grad-txt", type=Path, required=True)
    parser.add_argument("--workbook", type=Path, required=True)
    return parser.parse_args()


def clean_lines(path: Path) -> list[str]:
    out: list[str] = []
    for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw.strip()
        if not line:
            continue
        if any(line.startswith(prefix) for prefix in SKIP_PREFIXES):
            continue
        if line.startswith("— ") and line.endswith(" —"):
            continue
        out.append(line)
    return out


def split_records(lines: list[str]) -> list[tuple[int, list[str]]]:
    records: list[tuple[int, list[str]]] = []
    current_serial: int | None = None
    current_lines: list[str] = []

    for line in lines:
        match = SERIAL_RE.match(line)
        if match:
            serial = int(match.group(1))
            if current_serial is not None:
                records.append((current_serial, current_lines))
            current_serial = serial
            current_lines = []
            remainder = (match.group(2) or "").strip()
            if remainder:
                current_lines.append(remainder)
            continue

        if current_serial is not None:
            current_lines.append(line)

    if current_serial is not None:
        records.append((current_serial, current_lines))
    return records


def is_titleish(line: str) -> bool:
    return any(key in line for key in TITLE_KEYS)


def is_authorish(line: str) -> bool:
    if any(hint in line for hint in ORG_HINTS):
        return False
    if any(key in line for key in TITLE_KEYS):
        return False
    text = line.replace(" ", "")
    parts = [p for p in text.split("、") if p]
    if len(parts) >= 2 and all(len(p) <= 4 for p in parts):
        return True
    return False


def parse_record(lines: list[str]) -> tuple[str, str, str]:
    if not lines:
        return "", "", ""

    author_lines: list[str] = []
    idx = len(lines) - 1
    while idx >= 0:
        line = lines[idx]
        if is_authorish(line):
            author_lines.insert(0, line)
            idx -= 1
            continue
        # Allow a short author-continuation fragment like "昊、张志刚"
        if author_lines and len(line.replace(" ", "")) <= 12 and not any(k in line for k in TITLE_KEYS) and not any(h in line for h in ORG_HINTS):
            author_lines.insert(0, line)
            idx -= 1
            continue
        break

    upper = lines[: idx + 1]
    if not upper:
        return "", "", "".join(author_lines)

    school_end = 0
    seen_org = False
    for i, line in enumerate(upper):
        if any(hint in line for hint in ORG_HINTS):
            school_end = i + 1
            seen_org = True
            continue
        if seen_org and not is_titleish(line):
            school_end = i + 1
            continue
        if is_titleish(line):
            break
        if not seen_org and i == 0:
            school_end = 1
            continue
        break

    school_lines = upper[:school_end]
    title_lines = upper[school_end:]
    return "".join(school_lines), "".join(title_lines), "".join(author_lines)


def normalized(value: str) -> str:
    return value.replace(" ", "").replace("\u3000", "")


def main() -> None:
    args = parse_args()
    wb = load_workbook(args.workbook)
    ws = wb.active

    all_diffs: list[tuple[str, int, int, str, str, str]] = []

    for label, txt_path, offset in (
        ("undergrad", args.undergrad_txt, 1),
        ("grad", args.grad_txt, 216),
    ):
        records = split_records(clean_lines(txt_path))
        for serial, lines in records:
            school, title, authors = parse_record(lines)
            row = offset + serial
            current_title = str(ws.cell(row, 2).value or "")
            current_authors = str(ws.cell(row, 3).value or "")
            current_school = str(ws.cell(row, 4).value or "")

            if normalized(title) != normalized(current_title) or normalized(authors) != normalized(current_authors) or normalized(school) != normalized(current_school):
                all_diffs.append((label, serial, row, school, title, authors))

    print(f"DIFFS {len(all_diffs)}")
    for label, serial, row, school, title, authors in all_diffs:
        current = [ws.cell(row, c).value for c in range(1, 5)]
        print("ROW", row, "LABEL", label, "SERIAL", serial)
        print("EXPECTED", [school, title, authors])
        print("CURRENT", current)
        print("---")


if __name__ == "__main__":
    main()
