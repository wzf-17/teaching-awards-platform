from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def u(text: str) -> str:
    return text.encode("ascii").decode("unicode_escape")


ORG_HINTS = [
    u(r"\u5927\u5b66"),
    u(r"\u5b66\u9662"),
    u(r"\u7814\u7a76\u9662"),
    u(r"\u7814\u7a76\u6240"),
    u(r"\u4e2d\u5fc3"),
    u(r"\u57fa\u5730"),
    u(r"\u5b66\u6821"),
    u(r"\u516c\u53f8"),
    u(r"\u533b\u9662"),
]

TITLE_KEYS = [
    u(r"\u57f9\u517b"),
    u(r"\u5b9e\u8df5"),
    u(r"\u4f53\u7cfb"),
    u(r"\u6a21\u5f0f"),
    u(r"\u6784\u5efa"),
    u(r"\u521b\u65b0"),
    u(r"\u5bfc\u5411"),
    u(r"\u534f\u540c"),
    u(r"\u8d4b\u80fd"),
    u(r"\u80b2\u4eba"),
    u(r"\u6539\u9769"),
    u(r"\u63a2\u7d22"),
]

TRUNC_ENDS = set(
    [
        u(r"\u3001"),
        u(r"\uff0c"),
        u(r"\u4e2d"),
        u(r"\u897f"),
        u(r"\u7a7a"),
        u(r"\u7814"),
        u(r"\u5927"),
    ]
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    return parser.parse_args()


def is_title_org_like(title: str) -> bool:
    head = title[:24]
    return any(hint in head for hint in ORG_HINTS)


def is_org_truncated(org: str) -> bool:
    return bool(org) and org[-1] in TRUNC_ENDS


def is_org_has_title_words(org: str) -> bool:
    return any(key in org for key in TITLE_KEYS)


def is_short_title_long_org(title: str, org: str) -> bool:
    return len(title) <= 15 and len(org) >= 20


def main() -> None:
    args = parse_args()
    wb = load_workbook(args.workbook)
    ws = wb.active

    hits = []
    for row in range(2, ws.max_row + 1):
        title = str(ws.cell(row, 2).value or "")
        org = str(ws.cell(row, 4).value or "")
        issues = []
        if is_title_org_like(title):
            issues.append("title_org_like")
        if is_org_truncated(org):
            issues.append("org_truncated")
        if is_org_has_title_words(org):
            issues.append("org_has_title_words")
        if is_short_title_long_org(title, org):
            issues.append("short_title_long_org")
        if issues:
            hits.append((row, issues, [ws.cell(row, c).value for c in range(1, 5)]))

    print(f"HITS {len(hits)}")
    for row, issues, values in hits:
        safe = [str(v).encode("unicode_escape").decode("ascii") if v is not None else "None" for v in values]
        print(row, issues, safe)


if __name__ == "__main__":
    main()
