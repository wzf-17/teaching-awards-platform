from __future__ import annotations

import re
import subprocess
from dataclasses import dataclass
from pathlib import Path

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
from rapidocr_onnxruntime import RapidOCR


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = Path(__file__).resolve().parent
PDF_PATH = PROJECT_ROOT / "data" / "raw" / "provinces" / "2上海市" / "cda94fa0-312a-4bd8-830f-8f7d19efbb90.pdf"
WORK_DIR = PROJECT_ROOT / "data-pipeline" / "work" / "shanghai_awards_work_v2"
IMG_DIR = WORK_DIR / "images_300dpi"
OUTPUT_PATH = WORK_DIR / "上海市教学成果奖_527项_单表.xlsx"

COLS = {
    "serial": (322, 469),
    "unit": (469, 881),
    "title": (881, 1991),
    "people": (1991, 2952),
    "award": (2952, 3183),
}

SECTION_PAGE_RANGES = {
    "本科生": range(1, 30),
    "研究生": range(30, 37),
}

SINGLE_SURNAMES = set(
    "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦许何吕施张孔曹严华金魏陶姜戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳鲍史唐费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元卜顾孟平黄和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊纪舒屈项祝董梁杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林钟徐邱骆高夏蔡田樊胡凌霍虞万支柯昝管卢莫经房裘缪干解应宗丁宣贲邓郁单杭洪包诸左石崔吉龚程嵇邢裴陆荣翁荀羊於惠甄曲家封芮储靳汲邴糜松井段富巫乌焦巴弓牧隗山谷车侯宓蓬全郗班仰秋仲伊宫宁仇栾暴甘钭厉戎祖武符刘景詹束龙叶幸司韶郜黎蓟薄印宿白怀蒲台从鄂索咸籍赖卓蔺屠蒙池乔阴胥能苍双闻莘党翟谭贡劳逄姬申扶堵冉宰郦雍璩桑桂濮牛寿通边扈燕冀郏浦尚农温别庄晏柴瞿阎充慕连茹习宦艾鱼容向古易慎戈廖庾终暨居衡步都耿满弘匡国文寇广禄阙东殳沃利蔚越夔隆师巩聂晁勾敖融冷訾辛阚那简饶空曾毋沙乜养鞠须丰巢关蒯相查后荆红游竺权逯盖益桓公仉督岳帅缑亢况后有琴归海晋楚闫法汝鄢涂钦商牟佘佟伯赏墨哈谯笪年爱阳佴第五言福"
)
DOUBLE_SURNAMES = {
    "欧阳",
    "司马",
    "上官",
    "夏侯",
    "诸葛",
    "尉迟",
    "公孙",
    "司徒",
    "司空",
    "令狐",
    "皇甫",
    "长孙",
    "宇文",
    "慕容",
    "司寇",
    "南宫",
    "东方",
    "独孤",
    "东郭",
    "第五",
    "澹台",
    "公羊",
    "仲孙",
    "轩辕",
    "呼延",
    "端木",
    "巫马",
    "子车",
    "公西",
    "漆雕",
    "乐正",
    "壤驷",
    "公良",
    "拓跋",
}


@dataclass
class OcrItem:
    text: str
    x1: float
    x2: float
    y1: float
    y2: float

    @property
    def ym(self) -> float:
        return (self.y1 + self.y2) / 2


def section_from_page(page_no: int) -> str:
    if page_no in SECTION_PAGE_RANGES["本科生"]:
        return "本科生"
    if page_no in SECTION_PAGE_RANGES["研究生"]:
        return "研究生"
    return "成人教育"


def ensure_images() -> None:
    IMG_DIR.mkdir(parents=True, exist_ok=True)
    if any(IMG_DIR.glob("page-*.png")):
        return
    subprocess.run(
        [
            "pdftoppm.exe",
            "-png",
            "-r",
            "300",
            str(PDF_PATH),
            str(IMG_DIR / "page"),
        ],
        check=True,
    )


def normalize_text(text: str) -> str:
    text = str(text or "").replace("\u3000", " ").strip()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("一一", "——").replace("一 一", "——")
    return text.strip(" |")


def clean_award(text: str) -> str:
    text = normalize_text(text)
    if "特等" in text:
        return "特等"
    if "一等" in text:
        return "一等"
    if "二等" in text:
        return "二等"
    return text


def detect_row_boundaries(img_path: Path) -> list[int]:
    gray = np.array(Image.open(img_path).convert("L"))
    _thr, binary = cv2.threshold(gray, 220, 255, cv2.THRESH_BINARY_INV)
    horizontal = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (80, 1)),
    )
    horizontal = cv2.dilate(
        horizontal,
        cv2.getStructuringElement(cv2.MORPH_RECT, (5, 1)),
        iterations=1,
    )
    row_sum = (horizontal > 0).sum(axis=1)
    rows = np.where(row_sum > 800)[0]
    if len(rows) == 0:
        return []

    groups: list[list[int]] = [[int(rows[0])]]
    for raw in rows[1:]:
        row = int(raw)
        if row - groups[-1][-1] <= 20:
            groups[-1].append(row)
        else:
            groups.append([row])

    boundaries = [max(group, key=lambda idx: row_sum[idx]) for group in groups]
    return boundaries


def ocr_full_page(img: np.ndarray, engine: RapidOCR) -> list[OcrItem]:
    result, _ = engine(img)
    items: list[OcrItem] = []
    for box, text, _score in result or []:
        text = normalize_text(text)
        if not text:
            continue
        xs = [pt[0] for pt in box]
        ys = [pt[1] for pt in box]
        items.append(
            OcrItem(
                text=text,
                x1=min(xs),
                x2=max(xs),
                y1=min(ys),
                y2=max(ys),
            )
        )
    return items


def collect_column_text(items: list[OcrItem], left: int, right: int, top: int, bottom: int) -> str:
    chosen = [
        item
        for item in items
        if item.x1 >= left - 20 and item.x2 <= right + 20 and top <= item.ym <= bottom
    ]
    chosen.sort(key=lambda item: (round(item.ym / 10), item.x1))
    return normalize_text("".join(item.text for item in chosen))


def split_people_blob(blob: str) -> list[str]:
    blob = normalize_text(blob).replace(" ", "")
    blob = re.sub(r"[^\u4e00-\u9fff]", "", blob)
    if not blob:
        return []

    names: list[str] = []
    i = 0
    while i < len(blob):
        surname_len = 0
        if blob[i : i + 2] in DOUBLE_SURNAMES:
            surname_len = 2
        elif blob[i] in SINGLE_SURNAMES:
            surname_len = 1

        if surname_len == 0:
            i += 1
            continue

        remain = len(blob) - i - surname_len
        if remain <= 0:
            break

        given_len = 2 if remain >= 2 else 1
        name = blob[i : i + surname_len + given_len]
        if len(name) >= 2:
            names.append(name)
        i += surname_len + given_len

    deduped: list[str] = []
    seen: set[str] = set()
    for name in names:
        if name not in seen:
            deduped.append(name)
            seen.add(name)
    return deduped


def clean_name_text(text: str) -> str:
    text = normalize_text(text).replace(" ", "")
    return re.sub(r"[^\u4e00-\u9fff]", "", text)


def parse_standard_name_blob(blob: str) -> list[str] | None:
    if not blob:
        return []

    options: list[int] = []
    if blob[:2] in DOUBLE_SURNAMES:
        options.append(2)
    if blob[:1] in SINGLE_SURNAMES:
        options.append(1)

    for surname_len in options:
        for given_len in (2, 1):
            end = surname_len + given_len
            if end > len(blob):
                continue
            remainder = parse_standard_name_blob(blob[end:])
            if remainder is not None:
                return [blob[:end], *remainder]
    return None


def expand_name_token(name: str) -> list[str]:
    name = clean_name_text(name)
    if len(name) <= 4:
        return [name] if name else []

    parsed = parse_standard_name_blob(name)
    if parsed and len(parsed) > 1:
        return parsed
    return [name]


def detect_text_bands(binary: np.ndarray) -> list[tuple[int, int]]:
    row_sum = (binary > 0).sum(axis=1)
    rows = np.where(row_sum > 20)[0]
    if len(rows) == 0:
        return []

    groups: list[list[int]] = [[int(rows[0])]]
    for raw in rows[1:]:
        row = int(raw)
        if row - groups[-1][-1] <= 8:
            groups[-1].append(row)
        else:
            groups.append([row])

    bands: list[tuple[int, int]] = []
    for group in groups:
        top, bottom = group[0], group[-1]
        height = bottom - top + 1
        pixel_count = int(row_sum[top : bottom + 1].sum())
        if height < 12 or pixel_count < 1500:
            continue
        bands.append((top, bottom))
    return bands


def detect_name_segments(line_binary: np.ndarray) -> list[tuple[int, int]]:
    col_sum = (line_binary > 0).sum(axis=0)
    cols = np.where(col_sum > 3)[0]
    if len(cols) == 0:
        return []

    gap_threshold = max(20, int(line_binary.shape[0] * 0.22))
    groups: list[list[int]] = [[int(cols[0])]]
    for raw in cols[1:]:
        col = int(raw)
        if col - groups[-1][-1] <= gap_threshold:
            groups[-1].append(col)
        else:
            groups.append([col])

    min_width = max(25, int(line_binary.shape[0] * 0.30))
    return [
        (group[0], group[-1])
        for group in groups
        if group[-1] - group[0] + 1 >= min_width
    ]


def should_merge_name_fragments(prev_name: str, next_name: str) -> bool:
    if not prev_name or not next_name:
        return False
    return (
        len(prev_name) + len(next_name) <= 4
        and (len(prev_name) == 1 or len(next_name) == 1)
    )


def merge_line_edge_fragments(line_names: list[list[str]]) -> list[str]:
    merged: list[str] = []
    pending = ""

    for names in line_names:
        current = [name for name in names if len(name) >= 1]
        if not current:
            continue

        if pending:
            current[0] = pending + current[0]
            pending = ""

        if current and len(current[-1]) == 1:
            pending = current.pop()

        if (
            merged
            and current
            and should_merge_name_fragments(merged[-1], current[0])
        ):
            merged[-1] += current.pop(0)

        merged.extend(name for name in current if len(name) >= 2)

    if pending:
        if merged and should_merge_name_fragments(merged[-1], pending):
            merged[-1] += pending
        elif len(pending) >= 2:
            merged.append(pending)

    return merged


def extract_people(img: Image.Image, top: int, bottom: int, engine: RapidOCR) -> str:
    left, right = COLS["people"]
    crop = img.crop((left + 4, top, right - 4, bottom)).resize((right - left - 8, bottom - top))
    crop = crop.resize((crop.width * 2, crop.height * 2))
    crop_rgb = np.array(crop.convert("RGB"))
    gray = cv2.cvtColor(crop_rgb, cv2.COLOR_RGB2GRAY)
    _thr, binary = cv2.threshold(gray, 235, 255, cv2.THRESH_BINARY_INV)

    line_names: list[list[str]] = []
    name_patches: list[np.ndarray] = []
    line_ranges: list[tuple[int, int]] = []

    for y1, y2 in detect_text_bands(binary):
        line_top = max(0, y1 - 10)
        line_bottom = min(binary.shape[0], y2 + 11)
        line_binary = binary[line_top:line_bottom, :]
        segments = detect_name_segments(line_binary)
        if not segments:
            continue

        start_idx = len(name_patches)
        for x1, x2 in segments:
            patch_left = max(0, x1 - 12)
            patch_right = min(crop_rgb.shape[1], x2 + 13)
            name_patches.append(crop_rgb[line_top:line_bottom, patch_left:patch_right])
        line_ranges.append((start_idx, len(name_patches)))

    if name_patches:
        recognized, _elapsed = engine.text_recognizer(name_patches)
        for start_idx, end_idx in line_ranges:
            current: list[str] = []
            for text, _score in recognized[start_idx:end_idx]:
                current.extend(expand_name_token(text))
            if current:
                line_names.append(current)

    names = merge_line_edge_fragments(line_names)
    if names:
        return "、".join(names)

    result, _ = engine(crop_rgb)
    parts: list[tuple[float, float, str]] = []
    for box, text, _score in result or []:
        text = normalize_text(text)
        if not text:
            continue
        xs = [pt[0] for pt in box]
        ys = [pt[1] for pt in box]
        parts.append((min(xs), min(ys), text))
    parts.sort(key=lambda item: (round(item[1] / 10), item[0]))
    blob = "".join(item[2] for item in parts)
    return "、".join(split_people_blob(blob))


def extract_rows_for_page(img_path: Path, engine: RapidOCR) -> list[dict[str, str]]:
    page_no = int(re.search(r"page-(\d+)\.png", img_path.name).group(1))
    section = section_from_page(page_no)
    if section == "成人教育":
        return []

    boundaries = detect_row_boundaries(img_path)
    if len(boundaries) < 3:
        raise RuntimeError(f"failed to detect enough row boundaries: {img_path.name} -> {boundaries}")

    img = Image.open(img_path)
    img_np = np.array(img)
    page_items = ocr_full_page(img_np, engine)
    rows: list[dict[str, str]] = []

    for idx in range(1, len(boundaries) - 1):
        top = boundaries[idx] + 4
        bottom = boundaries[idx + 1] - 4
        unit = collect_column_text(page_items, *COLS["unit"], top, bottom)
        title = collect_column_text(page_items, *COLS["title"], top, bottom)
        award = clean_award(collect_column_text(page_items, *COLS["award"], top, bottom))
        people = extract_people(img, top, bottom, engine)

        rows.append(
            {
                "section": section,
                "title": title,
                "people": people,
                "unit": unit,
                "award": award,
            }
        )

    return rows


def build_workbook(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "上海市"
    headers = ["序号", "成果名称", "完成人", "完成单位", "获奖等级"]
    ws.append(headers)

    for idx, row in enumerate(rows, start=1):
        ws.append([idx, row["title"], row["people"], row["unit"], row["award"]])

    fill = PatternFill("solid", fgColor="4F81BD")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {"A": 10, "B": 58, "C": 54, "D": 30, "E": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        row[0].alignment = Alignment(horizontal="center", vertical="center")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)


def main() -> None:
    ensure_images()
    engine = RapidOCR()
    all_rows: list[dict[str, str]] = []

    for img_path in sorted(IMG_DIR.glob("page-*.png")):
        page_no = int(re.search(r"page-(\d+)\.png", img_path.name).group(1))
        section = section_from_page(page_no)
        if section == "成人教育":
            continue
        all_rows.extend(extract_rows_for_page(img_path, engine))

    undergrad_rows = [row for row in all_rows if row["section"] == "本科生"]
    graduate_rows = [row for row in all_rows if row["section"] == "研究生"]
    ordered_rows = undergrad_rows + graduate_rows

    print("本科生", len(undergrad_rows))
    print("研究生", len(graduate_rows))
    print("总计", len(ordered_rows))

    if len(undergrad_rows) != 429 or len(graduate_rows) != 98:
        raise RuntimeError(
            f"row count mismatch: undergrad={len(undergrad_rows)} graduate={len(graduate_rows)} total={len(ordered_rows)}"
        )

    build_workbook(ordered_rows)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
