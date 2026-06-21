from __future__ import annotations

import re
import subprocess
from collections import defaultdict
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
from rapidocr_onnxruntime import RapidOCR


PROJECT_ROOT = Path(__file__).resolve().parents[2]
ROOT = Path(__file__).resolve().parent
PDF_PATH = PROJECT_ROOT / "data" / "raw" / "provinces" / "2上海市" / "cda94fa0-312a-4bd8-830f-8f7d19efbb90.pdf"
WORK_DIR = PROJECT_ROOT / "data-pipeline" / "work" / "shanghai_awards_work"
IMG_DIR = WORK_DIR / "images_300dpi"
OUT_XLSX = WORK_DIR / "上海市教学成果奖_本科研究生整理.xlsx"

# Table borders from the 300 dpi render.
COL_LINES = [322, 469, 881, 1991, 2952, 3183]
ROW_THRESHOLD = 2200
PAGE_DPI = 300

AWARD_NAMES = ("特等", "一等", "二等")
SINGLE_SURNAME_CHARS = set(
    "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦许何吕施张孔曹严华金魏陶姜戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳鲍史唐费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元卜顾孟平黄和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊纪舒屈项祝董梁杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林钟徐邱骆高夏蔡田樊胡凌霍虞万支柯昝管卢莫经房裘缪干解应宗丁宣贲邓郁单杭洪包诸左石崔吉龚程嵇邢裴陆荣翁荀羊於惠甄曲家封芮储靳汲邴糜松井段富巫乌焦巴弓牧隗山谷车侯宓蓬全郗班仰秋仲伊宫宁仇栾暴甘钭厉戎祖武符刘景詹束龙叶幸司韶郜黎蓟薄印宿白怀蒲台从鄂索咸籍赖卓蔺屠蒙池乔阴胥能苍双闻莘党翟谭贡劳逄姬申扶堵冉宰郦雍璩桑桂濮牛寿通边扈燕冀郏浦尚农温别庄晏柴瞿阎充慕连茹习宦艾鱼容向古易慎戈廖庾终暨居衡步都耿满弘匡国文寇广禄阙东殳沃利蔚越夔隆师巩聂晁勾敖融冷訾辛阚那简饶空曾毋沙乜养鞠须丰巢关蒯相查后荆红游竺权逯盖益桓公仉督岳帅缑亢况后有琴归海晋楚闫法汝鄢涂钦商牟佘佟伯墨哈谯笪年爱阳佴第五言福"
)
MULTI_SURNAME_PARTS = {
    "欧阳",
    "司马",
    "上官",
    "夏侯",
    "诸葛",
    "尉迟",
    "公孙",
    "司徒",
    "司空",
    "令狐",
    "皇甫",
    "长孙",
    "宇文",
    "慕容",
    "司寇",
    "南宫",
    "东方",
    "独孤",
    "东郭",
    "第五",
    "澹台",
    "公羊",
    "仲孙",
    "轩辕",
    "呼延",
    "端木",
    "巫马",
    "子车",
    "公西",
    "漆雕",
    "乐正",
    "壤驷",
    "公良",
    "拓跋",
}


@dataclass
class OcrItem:
    text: str
    x1: float
    x2: float
    y1: float
    y2: float

    @property
    def ym(self) -> float:
        return (self.y1 + self.y2) / 2


def ensure_images() -> None:
    IMG_DIR.mkdir(parents=True, exist_ok=True)
    if any(IMG_DIR.glob("page-*.png")):
        return
    prefix = str(IMG_DIR / "page")
    subprocess.run(
        [
            "pdftoppm.exe",
            "-png",
            "-r",
            str(PAGE_DPI),
            str(PDF_PATH),
            prefix,
        ],
        check=True,
    )


def detect_row_lines(img: Image.Image) -> list[int]:
    gray = img.convert("L")
    arr = np.array(gray)
    mask = arr < 180
    row_dark = mask.sum(axis=1)
    raw = np.where(row_dark > ROW_THRESHOLD)[0]
    if len(raw) == 0:
        return []

    groups: list[tuple[int, int]] = []
    start = int(raw[0])
    prev = int(raw[0])
    for cur_raw in raw[1:]:
        cur = int(cur_raw)
        if cur == prev + 1:
            prev = cur
            continue
        groups.append((start, prev))
        start = prev = cur
    groups.append((start, prev))
    return [int((s + e) / 2) for s, e in groups]


def normalize_text(text: str) -> str:
    text = str(text or "").strip()
    text = text.replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text)
    text = text.replace("一一", "——").replace("一 一", "——")
    return text.strip(" |")


def is_surname_start(blob: str, idx: int) -> bool:
    if idx >= len(blob):
        return False
    if blob[idx : idx + 2] in MULTI_SURNAME_PARTS:
        return True
    return blob[idx] in SINGLE_SURNAME_CHARS


def split_name_blob(blob: str) -> list[str]:
    blob = normalize_text(blob).replace(" ", "")
    blob = re.sub(r"[^\u4e00-\u9fff]", "", blob)
    if not blob:
        return []

    @lru_cache(maxsize=None)
    def solve(idx: int) -> tuple[float, tuple[str, ...]]:
        if idx >= len(blob):
            return 0.0, ()

        candidates: list[tuple[float, tuple[str, ...]]] = []

        for surname_len in (2, 1):
            if surname_len == 2 and blob[idx : idx + 2] not in MULTI_SURNAME_PARTS:
                continue
            if surname_len == 1 and blob[idx] not in SINGLE_SURNAME_CHARS:
                continue

            for given_len, penalty in ((2, 0.0), (1, 0.22)):
                end = idx + surname_len + given_len
                if end > len(blob):
                    continue
                name = blob[idx:end]
                next_penalty = 0.0
                if end < len(blob) and not is_surname_start(blob, end):
                    next_penalty += 0.8
                rest_penalty, rest_names = solve(end)
                candidates.append((penalty + next_penalty + rest_penalty, (name,) + rest_names))

        rest_penalty, rest_names = solve(idx + 1)
        candidates.append((2.0 + rest_penalty, rest_names))
        return min(candidates, key=lambda item: (item[0], len(item[1])))

    _score, names = solve(0)
    cleaned: list[str] = []
    for name in names:
        if len(name) == 4 and name[:2] not in MULTI_SURNAME_PARTS:
            name = name[:3]
        if 2 <= len(name) <= 4:
            cleaned.append(name)
    return cleaned


def normalize_persons(entries: list[tuple[float, float, str]]) -> str:
    ordered = sorted(entries, key=lambda item: (round(item[1] / 8), item[0]))
    full_blob = "".join(raw for _x1, _y1, raw in ordered)
    names = split_name_blob(full_blob)

    deduped: list[str] = []
    seen: set[str] = set()
    for name in names:
        if name not in seen:
            deduped.append(name)
            seen.add(name)
    return "、".join(deduped)


def sort_items(items: list[OcrItem]) -> list[OcrItem]:
    return sorted(items, key=lambda item: (round(item.ym / 10), item.x1))


def collect_text(items: list[OcrItem]) -> str:
    ordered = sort_items(items)
    parts = [normalize_text(item.text) for item in ordered if normalize_text(item.text)]
    return normalize_text("".join(parts))


def clean_award(text: str) -> str:
    text = normalize_text(text)
    for award in AWARD_NAMES:
        if award in text:
            return award
    return text


def infer_section(page_name: str) -> str:
    page_no = int(re.search(r"page-(\d+)\.png", page_name).group(1))
    if page_no <= 29:
        return "本科教育"
    if page_no <= 36:
        return "研究生教育"
    return "成人教育"


def ocr_page(img_path: Path, engine: RapidOCR) -> list[OcrItem]:
    img = Image.open(img_path)
    result, _ = engine(np.array(img))
    items: list[OcrItem] = []
    for box, text, _score in result or []:
        text = normalize_text(text)
        if not text:
            continue
        xs = [p[0] for p in box]
        ys = [p[1] for p in box]
        items.append(OcrItem(text=text, x1=min(xs), x2=max(xs), y1=min(ys), y2=max(ys)))
    return items


def parse_pages() -> dict[str, list[dict[str, str]]]:
    ensure_images()
    engine = RapidOCR()
    pages = sorted(IMG_DIR.glob("page-*.png"))
    data: dict[str, list[dict[str, str]]] = defaultdict(list)
    section_serial = {"本科教育": 1, "研究生教育": 1}

    for img_path in pages:
        section = infer_section(img_path.name)
        if section == "成人教育":
            continue

        img = Image.open(img_path)
        row_lines = detect_row_lines(img)
        if len(row_lines) < 2:
            continue

        items = ocr_page(img_path, engine)
        header_bottom = row_lines[1]
        table_bottom = row_lines[-1]

        award_anchors = [
            item
            for item in items
            if item.x1 >= COL_LINES[4] - 20
            and header_bottom < item.ym < table_bottom + 60
            and clean_award(item.text) in AWARD_NAMES
        ]
        award_anchors.sort(key=lambda item: item.ym)
        if not award_anchors:
            continue

        for idx, anchor in enumerate(award_anchors):
            prev_mid = header_bottom if idx == 0 else (award_anchors[idx - 1].ym + anchor.ym) / 2
            next_mid = table_bottom if idx == len(award_anchors) - 1 else (anchor.ym + award_anchors[idx + 1].ym) / 2
            top = int(prev_mid) + 4
            bottom = int(next_mid) - 4
            if bottom <= top:
                continue

            row_items = [item for item in items if top <= item.ym <= bottom]
            if not row_items:
                continue

            unit_items = [item for item in row_items if item.x1 >= COL_LINES[1] - 10 and item.x2 <= COL_LINES[2] + 10]
            title_items = [item for item in row_items if item.x1 >= COL_LINES[2] - 10 and item.x2 <= COL_LINES[3] + 16]
            award_items = [item for item in row_items if item.x1 >= COL_LINES[4] - 16]

            person_crop = img.crop((COL_LINES[3] + 4, top - 10, COL_LINES[4] - 4, bottom + 10))
            person_result, _ = engine(np.array(person_crop))
            person_entries: list[tuple[float, float, str]] = []
            for box, text, _score in person_result or []:
                raw = normalize_text(text)
                if not raw:
                    continue
                xs = [p[0] for p in box]
                ys = [p[1] for p in box]
                person_entries.append((min(xs), min(ys), raw))

            title = collect_text(title_items)
            unit = collect_text(unit_items)
            award = clean_award(collect_text(award_items))
            persons = normalize_persons(person_entries)

            if not any([title, unit, persons, award]):
                continue

            serial = section_serial[section]
            section_serial[section] += 1
            data[section].append(
                {
                    "序号": serial,
                    "成果名称": title,
                    "完成人": persons,
                    "完成单位": unit,
                    "获奖等级": award,
                }
            )

    return data


def write_workbook(data: dict[str, list[dict[str, str]]]) -> None:
    wb = Workbook()
    headers = ["序号", "成果名称", "完成人", "完成单位", "获奖等级"]
    sheet_names = ["本科教育", "研究生教育"]

    for idx, sheet_name in enumerate(sheet_names):
        ws = wb.active if idx == 0 else wb.create_sheet(sheet_name)
        ws.title = sheet_name

        ws.append(headers)
        for row in data.get(sheet_name, []):
            ws.append([row[key] for key in headers])

        header_fill = PatternFill("solid", fgColor="4F81BD")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        widths = {"A": 10, "B": 56, "C": 54, "D": 30, "E": 12}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=2):
            row[0].alignment = Alignment(horizontal="center", vertical="center")
            row[1].alignment = Alignment(wrap_text=True, vertical="top")
            row[2].alignment = Alignment(wrap_text=True, vertical="top")
            row[3].alignment = Alignment(wrap_text=True, vertical="top")
            row[4].alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


def main() -> None:
    data = parse_pages()
    write_workbook(data)
    for sheet_name in ("本科教育", "研究生教育"):
        print(sheet_name, len(data.get(sheet_name, [])))
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
